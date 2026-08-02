from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from .models import (
    BuildPackage,
    BuildPackageItem,
    BuildPackageSection,
    ExtraOption,
    FoundationType,
    Material,
    Project,
    ProjectCategory,
    ProjectExtraOption,
    ProjectFoundation,
    ProjectOffer,
    ProjectPackageOverride,
    ProjectRoofCovering,
    RoofCovering,
)


class CatalogImportError(Exception):
    pass


def cell_to_str(value):
    return "" if value is None else str(value).strip()


def to_decimal(value):
    value = cell_to_str(value)
    if not value:
        return None
    try:
        return Decimal(value.replace(" ", "").replace(",", "."))
    except InvalidOperation as exc:
        raise CatalogImportError(f"Некорректное число: {value}") from exc


def to_int(value):
    value = cell_to_str(value)
    if not value:
        return None
    try:
        return int(float(value.replace(" ", "").replace(",", ".")))
    except ValueError as exc:
        raise CatalogImportError(f"Некорректное целое число: {value}") from exc


def to_bool(value, default=False):
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    value = cell_to_str(value).lower()
    if value in {"1", "true", "yes", "да", "д", "+", "активен"}:
        return True
    if value in {"0", "false", "no", "нет", "н", "-", ""}:
        return False
    raise CatalogImportError(f"Некорректное boolean-значение: {value}")


def get_sheet_rows(workbook, sheet_name, required=False):
    if sheet_name not in workbook.sheetnames:
        if required:
            raise CatalogImportError(f"В Excel-файле нет обязательного листа: {sheet_name}")
        return []
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    headers = [cell_to_str(cell) for cell in rows[0]]
    result = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        item = {header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
        item["_row_number"] = row_number
        result.append(item)
    return result


def get_category(value):
    value = cell_to_str(value)
    if not value:
        raise CatalogImportError("Категория проекта обязательна.")
    category = ProjectCategory.objects.filter(slug=value).first() or ProjectCategory.objects.filter(title__iexact=value).first()
    if not category:
        raise CatalogImportError(f"Категория не найдена: {value}")
    return category


def get_construction_type(value):
    value = cell_to_str(value).lower()
    values = {
        "брус": Project.ConstructionType.TIMBER,
        "timber": Project.ConstructionType.TIMBER,
        "каркас": Project.ConstructionType.FRAME,
        "frame": Project.ConstructionType.FRAME,
        "бревно": Project.ConstructionType.LOG,
        "log": Project.ConstructionType.LOG,
        "другое": Project.ConstructionType.OTHER,
        "other": Project.ConstructionType.OTHER,
    }
    if not value:
        return Project.ConstructionType.OTHER
    if value not in values:
        raise CatalogImportError(f"Неизвестный тип строительства: {value}")
    return values[value]


def _normalize(value):
    value = str(value or "").casefold().replace("×", "x").replace("х", "x")
    return re.sub(r"\s+", " ", value).strip()


def _stable_code(prefix, title):
    key = _normalize(title)
    known = {
        "обычный брус 150x150": "ordinary-150x150",
        "обычный брус 150x200": "ordinary-150x200",
        "профилированный брус 145x145": "profiled-145x145",
        "профилированный брус 145x195": "profiled-145x195",
        "брус камерной сушки 140x140": "dry-140x140",
        "брус камерной сушки 140x190": "dry-140x190",
        "свайный фундамент": "screw-piles",
        "свайно-винтовой фундамент": "screw-piles",
        "жб сваи (гост)": "reinforced-piles",
        "жб сваи гост": "reinforced-piles",
        "железобетонные сваи гост": "reinforced-piles",
        "металлочерепица": "metal-tile",
        "ондулин": "ondulin",
        "гибкая черепица": "flexible-shingles",
        "металлопрофиль": "metal-profile",
        "под усадку": "pod-usadku",
        'комплектация дома из бруса "под усадку"': "pod-usadku",
        "базовая комплектация": "pod-usadku",
    }
    if key in known:
        return known[key]
    return f"{prefix}-{hashlib.sha1(key.encode('utf-8')).hexdigest()[:10]}"


def get_or_create_material(group_title, title):
    group = _normalize(group_title)
    if "камер" in group or "суш" in group:
        kind = Material.Kind.DRY
    elif "проф" in group:
        kind = Material.Kind.PROFILED
    elif "брус" in group:
        kind = Material.Kind.REGULAR
    else:
        kind = Material.Kind.OTHER
    match = re.search(r"(\d{2,3})\s*x\s*(\d{2,3})", _normalize(title))
    material, _ = Material.objects.update_or_create(
        code=_stable_code("material", title),
        defaults={
            "kind": kind,
            "group_title": group_title or kind.label,
            "title": title,
            "section_width_mm": int(match.group(1)) if match else None,
            "section_height_mm": int(match.group(2)) if match else None,
            "is_active": True,
        },
    )
    return material


def get_or_create_build_package(title="Под усадку"):
    title = title or "Под усадку"
    package, _ = BuildPackage.objects.update_or_create(
        code=_stable_code("package", title),
        defaults={"title": title, "is_active": True},
    )
    return package


def get_or_create_foundation(title):
    code = _stable_code("foundation", title)
    existing = FoundationType.objects.filter(code=code).first()
    if existing:
        if existing.title != title and code != "reinforced-piles":
            existing.title = title
            existing.save(update_fields=["title"])
        return existing
    obj, _ = FoundationType.objects.update_or_create(
        code=code, defaults={"title": title, "is_active": True}
    )
    return obj


def get_or_create_roof_covering(title):
    obj, _ = RoofCovering.objects.update_or_create(
        code=_stable_code("roof", title), defaults={"title": title, "is_active": True}
    )
    return obj


def get_or_create_extra_option(title):
    obj, _ = ExtraOption.objects.update_or_create(
        code=_stable_code("extra", title), defaults={"title": title, "is_active": True}
    )
    return obj


def _sync_package_spec(project, package, spec):
    """Хранит общий состав один раз, а отличия — только как override проекта."""
    current = [
        {
            "title": section.title,
            "sort_order": section.sort_order,
            "items": [
                {"title": item.title, "value": item.value, "sort_order": item.sort_order}
                for item in section.items.all()
            ],
        }
        for section in package.sections.prefetch_related("items").all()
    ]
    if not current:
        for section_data in spec:
            section = BuildPackageSection.objects.create(
                package=package,
                title=section_data["title"],
                sort_order=section_data.get("sort_order", 0),
            )
            for item in section_data.get("items", []):
                BuildPackageItem.objects.create(
                    section=section,
                    title=item["title"],
                    value=item.get("value", ""),
                    sort_order=item.get("sort_order", 0),
                )
        ProjectPackageOverride.objects.filter(project=project, package=package).delete()
        return

    current_json = json.dumps(current, ensure_ascii=False, sort_keys=True)
    spec_json = json.dumps(spec, ensure_ascii=False, sort_keys=True)
    if current_json == spec_json:
        ProjectPackageOverride.objects.filter(project=project, package=package).delete()
    else:
        ProjectPackageOverride.objects.update_or_create(
            project=project,
            package=package,
            defaults={
                "sections": spec,
                "source_hash": hashlib.sha256(spec_json.encode("utf-8")).hexdigest(),
            },
        )


def import_catalog_excel(file_obj):
    workbook = load_workbook(file_obj, data_only=True)
    project_rows = get_sheet_rows(workbook, "projects", required=True)
    price_rows = get_sheet_rows(workbook, "price_options")
    addon_rows = get_sheet_rows(workbook, "addons")
    package_item_rows = get_sheet_rows(workbook, "package_items")

    result = {
        "projects_created": 0,
        "projects_updated": 0,
        "price_options_created": 0,
        "addons_created": 0,
        "package_items_created": 0,
    }
    imported_projects = {}

    with transaction.atomic():
        for row in project_rows:
            external_id = cell_to_str(row.get("external_id"))
            title = cell_to_str(row.get("title"))
            if not external_id or not title:
                raise CatalogImportError(f"Лист projects, строка {row['_row_number']}: external_id и title обязательны.")
            defaults = {
                "title": title,
                "category": get_category(row.get("category")),
                "construction_type": get_construction_type(row.get("construction_type")),
                "area": to_decimal(row.get("area")),
                "floors": to_decimal(row.get("floors")),
                "bedrooms": to_int(row.get("bedrooms")),
                "bathrooms": to_int(row.get("bathrooms")),
                "width": to_decimal(row.get("width")),
                "length": to_decimal(row.get("length")),
                "terrace_area": to_decimal(row.get("terrace_area")),
                "has_balcony": to_bool(row.get("has_balcony"), False),
                "has_porch": to_bool(row.get("has_porch"), False),
                "price_from": to_int(row.get("price_from")),  # legacy fallback only
                "build_days_from": to_int(row.get("build_days_from")),
                "build_days_to": to_int(row.get("build_days_to")),
                "short_description": cell_to_str(row.get("short_description")),
                "description": cell_to_str(row.get("description")),
                "seo_title": cell_to_str(row.get("seo_title")),
                "seo_description": cell_to_str(row.get("seo_description")),
                "is_active": to_bool(row.get("is_active"), True),
                "is_featured": to_bool(row.get("is_featured"), False),
                "sort_order": to_int(row.get("sort_order")) or 0,
            }
            slug = cell_to_str(row.get("slug"))
            if slug:
                defaults["slug"] = slugify(slug, allow_unicode=True)
            project, created = Project.objects.update_or_create(external_id=external_id, defaults=defaults)
            imported_projects[external_id] = project
            result["projects_created" if created else "projects_updated"] += 1

        ids = [project.id for project in imported_projects.values()]
        ProjectOffer.objects.filter(project_id__in=ids).delete()
        ProjectFoundation.objects.filter(project_id__in=ids).delete()
        ProjectRoofCovering.objects.filter(project_id__in=ids).delete()
        ProjectExtraOption.objects.filter(project_id__in=ids).delete()

        for row in price_rows:
            project = imported_projects.get(cell_to_str(row.get("project_external_id")))
            if not project:
                raise CatalogImportError(f"Лист price_options, строка {row['_row_number']}: проект не найден.")
            title = cell_to_str(row.get("title"))
            material = get_or_create_material(cell_to_str(row.get("group_title")), title)
            package = get_or_create_build_package(
                cell_to_str(row.get("package_title")) or cell_to_str(row.get("package_code")) or "Под усадку"
            )
            ProjectOffer.objects.create(
                project=project,
                material=material,
                build_package=package,
                base_price=to_int(row.get("price")),
                note=cell_to_str(row.get("note")),
                sort_order=to_int(row.get("sort_order")) or material.sort_order,
            )
            result["price_options_created"] += 1

        for row in addon_rows:
            project = imported_projects.get(cell_to_str(row.get("project_external_id")))
            if not project:
                raise CatalogImportError(f"Лист addons, строка {row['_row_number']}: проект не найден.")
            title = cell_to_str(row.get("title"))
            group = _normalize(row.get("group_title"))
            base_price = to_int(row.get("price"))
            description = cell_to_str(row.get("description"))
            sort_order = to_int(row.get("sort_order")) or 0
            if "фундамент" in group:
                obj = get_or_create_foundation(title)
                ProjectFoundation.objects.create(
                    project=project, foundation=obj, base_price_override=base_price,
                    description=description, sort_order=sort_order,
                )
            elif "кров" in group:
                obj = get_or_create_roof_covering(title)
                ProjectRoofCovering.objects.create(
                    project=project, covering=obj, base_price_override=base_price,
                    description=description, sort_order=sort_order,
                )
            else:
                obj = get_or_create_extra_option(title)
                ProjectExtraOption.objects.create(
                    project=project, option=obj, base_price_override=base_price,
                    description=description, sort_order=sort_order,
                )
            result["addons_created"] += 1

        # Старый Excel может повторять один и тот же состав комплектации для каждого проекта.
        # Сначала собираем спецификацию по (project, package), затем _sync_package_spec
        # оставит общий состав только один раз.
        grouped = defaultdict(lambda: defaultdict(list))
        for row in package_item_rows:
            project = imported_projects.get(cell_to_str(row.get("project_external_id")))
            if not project:
                raise CatalogImportError(f"Лист package_items, строка {row['_row_number']}: проект не найден.")
            package_title = cell_to_str(row.get("package_title")) or "Под усадку"
            section_title = cell_to_str(row.get("section_title")) or "Общее"
            grouped[(project.id, package_title)][section_title].append(
                {
                    "title": cell_to_str(row.get("title")),
                    "value": cell_to_str(row.get("value")),
                    "sort_order": to_int(row.get("sort_order")) or 0,
                }
            )
            result["package_items_created"] += 1

        for (project_id, package_title), sections in grouped.items():
            project = Project.objects.get(pk=project_id)
            package = get_or_create_build_package(package_title)
            spec = [
                {"title": title, "sort_order": index, "items": items}
                for index, (title, items) in enumerate(sections.items(), start=1)
            ]
            _sync_package_spec(project, package, spec)
            ProjectOffer.objects.filter(project=project, build_package__isnull=True).update(build_package=package)

    return result
